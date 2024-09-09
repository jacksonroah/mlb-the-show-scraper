import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import os

# Define the path to the workbook and the data folder
workbook_path = 'data/MLBTheShow24.xlsx'
data_folder = 'data'

# List of CSV files to update in the workbook
csv_files = ['team_performance.csv','player_performance_roah.csv', 'player_performance_lang.csv', 'game_log.csv']

# Load the existing workbook or create a new one if it doesn't exist
if os.path.exists(workbook_path):
    wb = openpyxl.load_workbook(workbook_path)
    print("Loaded existing workbook.")
else:
    wb = openpyxl.Workbook()
    print("Created new workbook.")

for csv_file in csv_files:
    file_path = os.path.join(data_folder, csv_file)
    if os.path.exists(file_path):
        df = pd.read_csv(file_path)
        sheet_name = csv_file.replace('.csv', '')
        
        # If the sheet already exists, remove it before adding a new one
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
            print(f"Removed existing sheet {sheet_name}.")
        
        # Add new sheet with updated data
        ws = wb.create_sheet(sheet_name)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        print(f"Updated sheet {sheet_name} with new data.")
    else:
        print(f"File {csv_file} does not exist in {data_folder}.")

# Save the workbook
wb.save(workbook_path)
print(f"Workbook saved to {workbook_path}")
